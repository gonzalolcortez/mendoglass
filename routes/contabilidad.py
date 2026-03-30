import io

from flask import Blueprint, render_template, request, send_file
from flask_login import login_required
from models import (
    db,
    MovimientoCaja,
    CUENTAS_CAJA,
    Cliente,
    Proveedor,
    Tecnico,
    ClienteCuentaCorrienteMovimiento,
    ProveedorCuentaCorrienteMovimiento,
    TecnicoCuentaCorrienteMovimiento,
    obtener_saldos_clientes,
    obtener_saldos_proveedores,
    obtener_saldos_tecnicos,
)
from datetime import datetime
from sqlalchemy import func, extract
from sqlalchemy.orm import joinedload

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

contabilidad_bp = Blueprint('contabilidad', __name__)


def _aplica_filtro_estado(saldo, estado):
    if estado == 'deben':
        return saldo > 0
    if estado == 'a_favor':
        return saldo < 0
    return abs(saldo) > 0.0001 or estado == 'todos'


def _filtrar_por_texto(nombre, q):
    if not q:
        return True
    return q.lower() in (nombre or '').lower()


def _build_cuentas_corrientes_context(args):
    q = (args.get('q') or '').strip()
    estado = (args.get('estado') or 'con_saldo').strip()
    entidad = (args.get('entidad') or 'todas').strip()

    clientes = Cliente.query.order_by(Cliente.apellido, Cliente.nombre).all()
    proveedores = Proveedor.query.order_by(Proveedor.apellido, Proveedor.nombre).all()
    tecnicos = Tecnico.query.order_by(Tecnico.nombre).all()
    saldos_clientes = obtener_saldos_clientes([cliente.id for cliente in clientes]) if clientes else {}
    saldos_proveedores = obtener_saldos_proveedores([proveedor.id for proveedor in proveedores]) if proveedores else {}
    saldos_tecnicos = obtener_saldos_tecnicos([tecnico.id for tecnico in tecnicos]) if tecnicos else {}

    for cliente in clientes:
        cliente.saldo_cc = saldos_clientes.get(cliente.id, 0.0)
    for proveedor in proveedores:
        proveedor.saldo_cc = saldos_proveedores.get(proveedor.id, 0.0)
    for tecnico in tecnicos:
        tecnico.saldo_cc = saldos_tecnicos.get(tecnico.id, 0.0)

    clientes_filtrados = [
        c for c in clientes
        if _filtrar_por_texto(c.nombre_completo, q)
        and _aplica_filtro_estado(c.saldo_cc, estado)
    ]
    proveedores_filtrados = [
        p for p in proveedores
        if _filtrar_por_texto(p.nombre_completo, q)
        and _aplica_filtro_estado(p.saldo_cc, estado)
    ]
    tecnicos_filtrados = [
        t for t in tecnicos
        if _filtrar_por_texto(t.nombre_display, q)
        and _aplica_filtro_estado(t.saldo_cc, estado)
    ]

    if entidad == 'clientes':
        proveedores_filtrados = []
        tecnicos_filtrados = []
    elif entidad == 'proveedores':
        clientes_filtrados = []
        tecnicos_filtrados = []
    elif entidad == 'tecnicos':
        clientes_filtrados = []
        proveedores_filtrados = []

    total_clientes_deuda = sum(max(c.saldo_cc, 0.0) for c in clientes_filtrados)
    total_clientes_a_favor = sum(abs(min(c.saldo_cc, 0.0)) for c in clientes_filtrados)
    total_proveedores_deuda = sum(max(p.saldo_cc, 0.0) for p in proveedores_filtrados)
    total_proveedores_a_favor = sum(abs(min(p.saldo_cc, 0.0)) for p in proveedores_filtrados)
    total_tecnicos_deuda = sum(max(t.saldo_cc, 0.0) for t in tecnicos_filtrados)
    total_tecnicos_a_favor = sum(abs(min(t.saldo_cc, 0.0)) for t in tecnicos_filtrados)

    cliente_ids = {c.id for c in clientes_filtrados}
    proveedor_ids = {p.id for p in proveedores_filtrados}
    tecnico_ids = {t.id for t in tecnicos_filtrados}

    recientes = []
    if entidad in ('todas', 'clientes'):
        movimientos_cli_query = (
            ClienteCuentaCorrienteMovimiento.query
            .options(joinedload(ClienteCuentaCorrienteMovimiento.cliente))
            .order_by(ClienteCuentaCorrienteMovimiento.fecha.desc())
        )
        if cliente_ids:
            movimientos_cli_query = movimientos_cli_query.filter(
                ClienteCuentaCorrienteMovimiento.cliente_id.in_(cliente_ids)
            )
        movimientos_cli = movimientos_cli_query.limit(80).all() if cliente_ids else []
        for m in movimientos_cli:
            recientes.append({
                'fecha': m.fecha,
                'entidad_tipo': 'cliente',
                'entidad_nombre': m.cliente.nombre_completo if m.cliente else f'Cliente #{m.cliente_id}',
                'entidad_id': m.cliente_id,
                'tipo': m.tipo,
                'concepto': m.concepto,
                'monto': m.monto,
            })
    if entidad in ('todas', 'proveedores'):
        movimientos_prov_query = (
            ProveedorCuentaCorrienteMovimiento.query
            .options(joinedload(ProveedorCuentaCorrienteMovimiento.proveedor))
            .order_by(ProveedorCuentaCorrienteMovimiento.fecha.desc())
        )
        if proveedor_ids:
            movimientos_prov_query = movimientos_prov_query.filter(
                ProveedorCuentaCorrienteMovimiento.proveedor_id.in_(proveedor_ids)
            )
        movimientos_prov = movimientos_prov_query.limit(80).all() if proveedor_ids else []
        for m in movimientos_prov:
            recientes.append({
                'fecha': m.fecha,
                'entidad_tipo': 'proveedor',
                'entidad_nombre': m.proveedor.nombre_completo if m.proveedor else f'Proveedor #{m.proveedor_id}',
                'entidad_id': m.proveedor_id,
                'tipo': m.tipo,
                'concepto': m.concepto,
                'monto': m.monto,
            })
    if entidad in ('todas', 'tecnicos'):
        movimientos_tec_query = (
            TecnicoCuentaCorrienteMovimiento.query
            .options(joinedload(TecnicoCuentaCorrienteMovimiento.tecnico))
            .order_by(TecnicoCuentaCorrienteMovimiento.fecha.desc())
        )
        if tecnico_ids:
            movimientos_tec_query = movimientos_tec_query.filter(
                TecnicoCuentaCorrienteMovimiento.tecnico_id.in_(tecnico_ids)
            )
        movimientos_tec = movimientos_tec_query.limit(80).all() if tecnico_ids else []
        for m in movimientos_tec:
            recientes.append({
                'fecha': m.fecha,
                'entidad_tipo': 'tecnico',
                'entidad_nombre': m.tecnico.nombre_display if m.tecnico else f'Técnico #{m.tecnico_id}',
                'entidad_id': m.tecnico_id,
                'tipo': m.tipo,
                'concepto': m.concepto,
                'monto': m.monto,
            })

    recientes.sort(key=lambda x: x['fecha'] or datetime.min, reverse=True)
    recientes = recientes[:100]

    return {
        'q': q,
        'estado': estado,
        'entidad': entidad,
        'clientes_con_saldo': clientes_filtrados,
        'proveedores_con_saldo': proveedores_filtrados,
        'total_clientes_deuda': total_clientes_deuda,
        'total_clientes_a_favor': total_clientes_a_favor,
        'total_proveedores_deuda': total_proveedores_deuda,
        'total_proveedores_a_favor': total_proveedores_a_favor,
        'tecnicos_con_saldo': tecnicos_filtrados,
        'total_tecnicos_deuda': total_tecnicos_deuda,
        'total_tecnicos_a_favor': total_tecnicos_a_favor,
        'recientes': recientes,
    }


def _exportar_cuentas_corrientes_xlsx(contexto):
    workbook = Workbook()
    resumen = workbook.active
    resumen.title = 'Resumen'
    resumen.append(['Concepto', 'Monto'])
    resumen.append(['Clientes que deben', contexto['total_clientes_deuda']])
    resumen.append(['Clientes a favor', contexto['total_clientes_a_favor']])
    resumen.append(['Deuda con proveedores', contexto['total_proveedores_deuda']])
    resumen.append(['Saldo a favor proveedores', contexto['total_proveedores_a_favor']])

    ws_clientes = workbook.create_sheet('Clientes')
    ws_clientes.append(['ID', 'Cliente', 'Estado', 'Saldo'])
    for cliente in contexto['clientes_con_saldo']:
        estado = 'Debe' if cliente.saldo_cc > 0 else 'A favor'
        ws_clientes.append([cliente.id, cliente.nombre_completo, estado, abs(cliente.saldo_cc)])

    ws_proveedores = workbook.create_sheet('Proveedores')
    ws_proveedores.append(['ID', 'Proveedor', 'Estado', 'Saldo'])
    for proveedor in contexto['proveedores_con_saldo']:
        estado = 'Debemos' if proveedor.saldo_cc > 0 else 'A favor'
        ws_proveedores.append([proveedor.id, proveedor.nombre_completo, estado, abs(proveedor.saldo_cc)])

    ws_tecnicos = workbook.create_sheet('Tecnicos')
    ws_tecnicos.append(['ID', 'Tecnico', 'Estado', 'Saldo'])
    for tecnico in contexto['tecnicos_con_saldo']:
        estado = 'Debe' if tecnico.saldo_cc > 0 else 'A favor'
        ws_tecnicos.append([tecnico.id, tecnico.nombre_display, estado, abs(tecnico.saldo_cc)])

    ws_movs = workbook.create_sheet('Movimientos')
    ws_movs.append(['Fecha', 'Entidad', 'Nombre', 'Tipo', 'Concepto', 'Monto'])
    for mov in contexto['recientes']:
        ws_movs.append([
            mov['fecha'].strftime('%d/%m/%Y %H:%M') if mov['fecha'] else '',
            mov['entidad_tipo'],
            mov['entidad_nombre'],
            mov['tipo'],
            mov['concepto'],
            mov['monto'],
        ])

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _exportar_cuentas_corrientes_pdf(contexto):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    story = [Paragraph('Cuentas Corrientes', styles['Title']), Spacer(1, 12)]

    resumen_data = [
        ['Concepto', 'Monto'],
        ['Clientes que deben', f"${contexto['total_clientes_deuda']:.2f}"],
        ['Clientes a favor', f"${contexto['total_clientes_a_favor']:.2f}"],
        ['Deuda con proveedores', f"${contexto['total_proveedores_deuda']:.2f}"],
        ['Saldo a favor proveedores', f"${contexto['total_proveedores_a_favor']:.2f}"],
        ['Tecnicos que deben', f"${contexto['total_tecnicos_deuda']:.2f}"],
        ['Tecnicos a favor', f"${contexto['total_tecnicos_a_favor']:.2f}"],
    ]
    tabla_resumen = Table(resumen_data, hAlign='LEFT', colWidths=[240, 120])
    tabla_resumen.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dbe5f1')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    ]))
    story.extend([tabla_resumen, Spacer(1, 16)])

    if contexto['clientes_con_saldo']:
        story.append(Paragraph('Clientes', styles['Heading2']))
        data = [['Cliente', 'Estado', 'Saldo']] + [
            [c.nombre_completo, 'Debe' if c.saldo_cc > 0 else 'A favor', f"${abs(c.saldo_cc):.2f}"]
            for c in contexto['clientes_con_saldo']
        ]
        tabla = Table(data, hAlign='LEFT', colWidths=[260, 120, 120])
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#eef3f8')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ]))
        story.extend([tabla, Spacer(1, 16)])

    if contexto['proveedores_con_saldo']:
        story.append(Paragraph('Proveedores', styles['Heading2']))
        data = [['Proveedor', 'Estado', 'Saldo']] + [
            [p.nombre_completo, 'Debemos' if p.saldo_cc > 0 else 'A favor', f"${abs(p.saldo_cc):.2f}"]
            for p in contexto['proveedores_con_saldo']
        ]
        tabla = Table(data, hAlign='LEFT', colWidths=[260, 120, 120])
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#eef3f8')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ]))
        story.extend([tabla, Spacer(1, 16)])

    if contexto['tecnicos_con_saldo']:
        story.append(Paragraph('Tecnicos', styles['Heading2']))
        data = [['Tecnico', 'Estado', 'Saldo']] + [
            [t.nombre_display, 'Debe' if t.saldo_cc > 0 else 'A favor', f"${abs(t.saldo_cc):.2f}"]
            for t in contexto['tecnicos_con_saldo']
        ]
        tabla = Table(data, hAlign='LEFT', colWidths=[260, 120, 120])
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#eef3f8')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ]))
        story.extend([tabla, Spacer(1, 16)])

    if contexto['recientes']:
        story.append(Paragraph('Movimientos recientes', styles['Heading2']))
        data = [['Fecha', 'Entidad', 'Nombre', 'Tipo', 'Concepto', 'Monto']] + [
            [
                m['fecha'].strftime('%d/%m/%Y %H:%M') if m['fecha'] else '',
                m['entidad_tipo'].title(),
                m['entidad_nombre'],
                m['tipo'].title(),
                m['concepto'],
                f"${m['monto']:.2f}",
            ]
            for m in contexto['recientes'][:40]
        ]
        tabla = Table(data, hAlign='LEFT', colWidths=[90, 80, 160, 70, 240, 80])
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#eef3f8')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]))
        story.append(tabla)

    doc.build(story)
    buffer.seek(0)
    return buffer


@contabilidad_bp.route('/')
@login_required
def index():
    anio = request.args.get('anio', datetime.now().year, type=int)

    # Ingresos y egresos por mes del año seleccionado
    meses_data = []
    for mes in range(1, 13):
        ingresos = db.session.query(func.sum(MovimientoCaja.monto)).filter(
            MovimientoCaja.tipo == 'ingreso',
            extract('year', MovimientoCaja.fecha) == anio,
            extract('month', MovimientoCaja.fecha) == mes,
        ).scalar() or 0.0

        egresos = db.session.query(func.sum(MovimientoCaja.monto)).filter(
            MovimientoCaja.tipo == 'egreso',
            extract('year', MovimientoCaja.fecha) == anio,
            extract('month', MovimientoCaja.fecha) == mes,
        ).scalar() or 0.0

        meses_data.append({
            'mes': mes,
            'nombre': _nombre_mes(mes),
            'ingresos': ingresos,
            'egresos': egresos,
            'ganancia': ingresos - egresos,
        })

    total_ingresos = sum(m['ingresos'] for m in meses_data)
    total_egresos = sum(m['egresos'] for m in meses_data)
    total_ganancia = total_ingresos - total_egresos

    # Años disponibles
    anios = db.session.query(
        extract('year', MovimientoCaja.fecha).label('anio')
    ).distinct().order_by('anio').all()
    anios = [int(a.anio) for a in anios] or [datetime.now().year]
    if datetime.now().year not in anios:
        anios.append(datetime.now().year)
    anios.sort(reverse=True)

    # Saldo acumulado por cuenta (todos los movimientos)
    # Saldo acumulado por cuenta (todos los movimientos) — dos queries con GROUP BY
    rows_ing = db.session.query(
        MovimientoCaja.cuenta, func.sum(MovimientoCaja.monto)
    ).filter_by(tipo='ingreso').group_by(MovimientoCaja.cuenta).all()
    rows_egr = db.session.query(
        MovimientoCaja.cuenta, func.sum(MovimientoCaja.monto)
    ).filter_by(tipo='egreso').group_by(MovimientoCaja.cuenta).all()

    ing_por_cuenta = {cuenta: total for cuenta, total in rows_ing}
    egr_por_cuenta = {cuenta: total for cuenta, total in rows_egr}

    saldos_cuentas = []
    for cuenta_val, cuenta_label in CUENTAS_CAJA:
        ing = ing_por_cuenta.get(cuenta_val, 0.0)
        egr = egr_por_cuenta.get(cuenta_val, 0.0)
        saldos_cuentas.append({
            'cuenta': cuenta_val,
            'nombre': cuenta_label,
            'ingresos': ing,
            'egresos': egr,
            'saldo': ing - egr,
        })

    return render_template('contabilidad/index.html',
                           meses_data=meses_data,
                           total_ingresos=total_ingresos,
                           total_egresos=total_egresos,
                           total_ganancia=total_ganancia,
                           anio=anio,
                           anios=anios,
                           saldos_cuentas=saldos_cuentas)


@contabilidad_bp.route('/cuenta/<cuenta>')
@login_required
def detalle_cuenta(cuenta):
    cuentas_dict = dict(CUENTAS_CAJA)
    nombre_cuenta = cuentas_dict.get(cuenta, cuenta)

    movimientos = MovimientoCaja.query.filter_by(cuenta=cuenta)\
        .order_by(MovimientoCaja.fecha.asc()).all()

    # Calcular saldo progresivo
    saldo = 0.0
    movimientos_con_saldo = []
    for m in movimientos:
        if m.tipo == 'ingreso':
            saldo += m.monto
        else:
            saldo -= m.monto
        movimientos_con_saldo.append({
            'mov': m,
            'ingreso': m.monto if m.tipo == 'ingreso' else None,
            'egreso': m.monto if m.tipo == 'egreso' else None,
            'saldo': saldo,
        })

    total_ingresos = sum(r['ingreso'] for r in movimientos_con_saldo if r['ingreso'])
    total_egresos = sum(r['egreso'] for r in movimientos_con_saldo if r['egreso'])

    return render_template('contabilidad/cuenta.html',
                           cuenta=cuenta,
                           nombre_cuenta=nombre_cuenta,
                           movimientos_con_saldo=movimientos_con_saldo,
                           total_ingresos=total_ingresos,
                           total_egresos=total_egresos,
                           saldo_final=saldo)


@contabilidad_bp.route('/cuentas-corrientes')
@login_required
def cuentas_corrientes():
    contexto = _build_cuentas_corrientes_context(request.args)
    return render_template('contabilidad/cuentas_corrientes.html', **contexto)


@contabilidad_bp.route('/cuentas-corrientes/export/<formato>')
@login_required
def exportar_cuentas_corrientes(formato):
    contexto = _build_cuentas_corrientes_context(request.args)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    if formato == 'xlsx':
        buffer = _exportar_cuentas_corrientes_xlsx(contexto)
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'cuentas_corrientes_{timestamp}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    if formato == 'pdf':
        buffer = _exportar_cuentas_corrientes_pdf(contexto)
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'cuentas_corrientes_{timestamp}.pdf',
            mimetype='application/pdf',
        )

    return ('Formato no soportado', 400)


def _nombre_mes(n):
    nombres = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
               'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    return nombres[n - 1]
